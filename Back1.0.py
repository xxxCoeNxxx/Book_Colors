from tkinter import *
from openpyxl import Workbook, load_workbook

# Cargar el libro existente o crear uno nuevo si no existe
try:
    book = load_workbook("datos_libros.xlsx")
except FileNotFoundError:
    book = Workbook()

sheet = book.active

numero_libro = 1

class Figuras(Tk):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.geometry("424x624")

        self.label_text = "Libro  {}".format(numero_libro)
        self.label = Label(self, text=self.label_text, font=("Helvetica", 16))
        self.label.place(x=10, y=10)

        self.create_canvas(0, 100, self.cuadrado_rojo_function, "red", "Rojo")
        self.create_canvas(210, 100, self.cuadrado_gris_function, "grey", "Gris")
        self.create_canvas(0, 310, self.cuadrado_ninguno_function, "green", "Ninguno")
        self.create_canvas(210, 310, self.cuadrado_prestado_function, "blue", "Prestado")
        self.cuadrado_back_button()
        self.cuadrado_next_button()

    def create_canvas(self, x, y, command, color, opcion):
        canvas = Canvas(self, width=210, height=210, bg="black")
        canvas.place(x=x, y=y)

        rectangulo = canvas.create_rectangle(10, 10, 200, 200, width=4, fill=color)
        canvas.tag_bind(rectangulo, '<Button-1>', lambda event, command=command, opcion=opcion: command(opcion))

    def cuadrado_rojo_function(self, opcion):
        global numero_libro
        self.guardar_en_excel(numero_libro, opcion, sobrescribir=True)
        self.ordenar_excel()
        numero_libro += 1
        self.actualizar_label()
        print(opcion, numero_libro)

    def cuadrado_gris_function(self, opcion):
        global numero_libro
        self.guardar_en_excel(numero_libro, opcion, sobrescribir=True)
        self.ordenar_excel()
        numero_libro += 1
        self.actualizar_label()
        print(opcion, numero_libro)

    def cuadrado_ninguno_function(self, opcion):
        global numero_libro
        self.guardar_en_excel(numero_libro, opcion, sobrescribir=True)
        self.ordenar_excel()
        numero_libro += 1
        self.actualizar_label()
        print(opcion, numero_libro)

    def cuadrado_prestado_function(self, opcion):
        global numero_libro
        self.guardar_en_excel(numero_libro, opcion, sobrescribir=True)
        self.ordenar_excel()
        numero_libro += 1
        self.actualizar_label()
        print(opcion, numero_libro)

    def cuadrado_back_button(self):
        button = Button(self, text="Anterior", command=self.cuadrado_back_function)
        button.place(x=0, y=515, width=100, height=200)

    def cuadrado_back_function(self, sobrescribir=False):
        global numero_libro
        self.ordenar_excel()
        numero_libro -= 1
        self.actualizar_label()
        print("Libro anterior.")

    def cuadrado_next_button(self):
        button = Button(self, text="Siguiente", command=self.cuadrado_next_function)
        button.place(x=325, y=315, width=100, height=200)

    def cuadrado_next_function(self):
        global numero_libro
        self.ordenar_excel()
        numero_libro += 1
        self.actualizar_label()
        print("Siguiente libro.")

    def actualizar_label(self):
        self.label_text = "Libro  {}".format(numero_libro)
        self.label.config(text=self.label_text)

    def guardar_en_excel(self, libro, opcion, sobrescribir=False):
        # Buscar si el valor de numero_libro ya está en la tabla
        encontrado = False

        for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row):
            if row[0].value == libro:
                # Si se encuentra, actualizar los valores y salir del bucle
                encontrado = True
                if sobrescribir:
                    sheet.cell(row=row[0].row, column=2, value=opcion)
                break

        # Si no se encuentra o no se debe sobrescribir, agregar una nueva fila
        if not encontrado or not sobrescribir:
            sheet.append([libro, opcion])

        book.save("datos_libros.xlsx")

    def ordenar_excel(self):
        # Cargar el libro existente
        book = load_workbook("datos_libros.xlsx")

        # Seleccionar la hoja activa
        sheet = book.active

        # Obtener todas las filas excluyendo la primera (encabezados)
        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        # Ordenar las filas por la primera columna (Libro)
        sorted_rows = sorted(rows, key=lambda x: x[0])

        # Limpiar la hoja antes de escribir las filas ordenadas
        sheet.delete_rows(2, sheet.max_row)

        # Escribir las filas ordenadas de nuevo en la hoja
        for row in sorted_rows:
            sheet.append(row)

        # Guardar el libro con el nuevo orden
        book.save("datos_libros_ordenados.xlsx")

if __name__ == "__main__":
    app = Figuras()
    app.mainloop()
